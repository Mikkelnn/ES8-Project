import numpy as np
from openpyxl import load_workbook

min_range = 2
max_range = 21931
seed = 2651


def main():
    # Use the found seed
    np.random.seed(seed)
    randNum = np.random.randint(min_range, max_range + 1)
    
    # Load the Excel file
    try:
        wb = load_workbook("drenge_navne.xlsx", data_only=True)
        ws = wb.active
        
        # Read the name from column A at the row index (randNum)
        name = str(ws[f"A{randNum}"].value)
        print(f"Random name from the list of approved names in Denmark: {name}")
    except Exception as e:
        print(f"Error reading Excel file: {e}")



if __name__ == "__main__":
    main()